import pyesedb
from datetime import datetime,timedelta
import sys
import struct
import re
import openpyxl
from openpyxl.writer.write_only import WriteOnlyCell
import argparse
import warnings
import hashlib
import random


def BinarySIDtoStringSID(sid):
  #Source: https://github.com/google/grr/blob/master/grr/parsers/wmi_parser.py
  """Converts a binary SID to its string representation.
  https://msdn.microsoft.com/en-us/library/windows/desktop/aa379597.aspx
  The byte representation of an SID is as follows:
    Offset  Length  Description
    00      01      revision
    01      01      sub-authority count
    02      06      authority (big endian)
    08      04      subauthority #1 (little endian)
    0b      04      subauthority #2 (little endian)
    ...
  Args:
    sid: A byte array.
  Returns:
    SID in string form.
  Raises:
    ValueError: If the binary SID is malformed.
  """
  if not sid:
    return ""
  str_sid_components = [ord(sid[0])]
  # Now decode the 48-byte portion
  if len(sid) >= 8:
    subauthority_count = ord(sid[1])
    identifier_authority = struct.unpack(">H", sid[2:4])[0]
    identifier_authority <<= 32
    identifier_authority |= struct.unpack(">L", sid[4:8])[0]
    str_sid_components.append(identifier_authority)
    start = 8
    for i in range(subauthority_count):
      authority = sid[start:start + 4]
      if not authority:
        break
      if len(authority) < 4:
        raise ValueError("In binary SID '%s', component %d has been truncated. "
                         "Expected 4 bytes, found %d: (%s)",
                         ",".join([str(ord(c)) for c in sid]), i,
                         len(authority), authority)
      str_sid_components.append(struct.unpack("<L", authority)[0])
      start += 4
  return "S-%s" % ("-".join([str(x) for x in str_sid_components]))

def ole_timestamp(binblob,timeformat="%Y-%m-%d %H:%M:%S"):
    #converts a hex encoded OLE time stamp to a time string
    ts = struct.unpack("<d",binblob)[0]
    dt = datetime(1899,12,30,0,0,0) + timedelta(days=ts)
    return  dt.strftime(timeformat)

def file_timestamp(binblob,timeformat="%Y-%m-%d %H:%M:%S"):
    #converts a hex encoded windows file time stamp to a time string
    dt = datetime(1601,1,1,0,0,0) + timedelta(microseconds=binblob/10)
    return  dt.strftime(timeformat)

def blob_to_string(binblob):
    try:
        if re.match('^[ -~]+\x00?$', binblob.decode("utf-16-le")):
            binblob = binblob.decode("utf-16-le").strip("\x00")
        elif re.match('^[ -~]+\x00?$', binblob.decode("utf-16-be")):
            binblob = binblob.decode("utf-16-be").strip("\x00")
        elif re.match('^[ -~]+\x00?$', binblob.encode("ascii","ignore")):
            binblob = binblob.encode("ascii","ignore").strip("\x00")
    except:
        binblob = "" if not binblob else binblob.encode("HEX")
    return binblob

def load_interfaces(reg_file):
    from Registry import Registry
    import struct
    try:
        reg_handle = Registry.Registry("SOFTWARE")
    except Exception as e:
        print "I could not open the specified SOFTWARE registry key. It is usually located in \Windows\system32\config.  This is an optional value.  If you cant find it just dont provide one."
        print "Error : ", str(e)
        sys.exit(1)
    int_keys = reg_handle.open('Microsoft\\WlanSvc\\Interfaces')
    profile_lookup = {}
    for eachinterface in int_keys.subkeys():
        if len(eachinterface.subkeys())==0:
            continue
        for eachprofile in eachinterface.subkey("Profiles").subkeys():
            profileid = [x.value() for x in eachprofile.values() if x.name()=="ProfileIndex"][0]
            metadata = eachprofile.subkey("MetaData").values()
            for eachvalue in metadata:
                if eachvalue.name()=="Channel Hints":
                    channelhintraw = eachvalue.value()
                    hintlength = struct.unpack("I", channelhintraw[0:4])[0]
                    name = channelhintraw[4:hintlength+4] 
                    profile_lookup[str(profileid)] = name
    return profile_lookup

def load_lookups(database):
    id_lookup = {}
    #Note columns  0 = Type, 1 = Index, 2 = Value
    lookup_table = database.get_table_by_name('SruDbIdMapTable')
    column_lookup = dict([(x.name,index) for index,x in enumerate(lookup_table.columns)])
    for rec_entry_num in range(lookup_table.number_of_records):
        if smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdType'])==0:
            #proc_blob = 'None' if not rec_entry.get_value_data(2) else unicode(rec_entry.get_value_data(2).decode("hex"),'utf-16-le').strip("\x00")
            proc_blob = smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdBlob'])
            if not proc_blob == "Empty":
                proc_blob = blob_to_string(proc_blob)
            id_lookup[smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdIndex'])] = proc_blob
        elif smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdType'])==1:
            #id_lookup[rec_entry.get_value_data_as_integer(1)] = unicode(rec_entry.get_value_data(2).decode("hex"),'utf-16-le').strip("\x00")
            proc_blob = smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdBlob'])
            if not proc_blob == "Empty":
                proc_blob = blob_to_string(proc_blob)
            id_lookup[smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdIndex'])] = proc_blob
        elif smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdType'])==2:
            #d_lookup[rec_entry.get_value_data_as_integer(1)] = unicode(rec_entry.get_value_data(2).decode("hex"),'utf-16-le').strip("\x00")
            proc_blob = smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdBlob'])
            if not proc_blob == "Empty":
                proc_blob = blob_to_string(proc_blob)
            id_lookup[smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdIndex'])] = proc_blob
        elif smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdType'])==3:
            user_blob = BinarySIDtoStringSID(smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdBlob']))
            id_lookup[smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdIndex'])] = user_blob
        else:
            print "unknown entry type in IdMapTable"
    return id_lookup
                #print rec_entry

def smart_retrieve(ese_table, ese_record_num, column_number):
	ese_column_types = {0: 'NULL', 1: 'BOOLEAN', 2: 'INTEGER_8BIT_UNSIGNED', 3: 'INTEGER_16BIT_SIGNED', 4: 'INTEGER_32BIT_SIGNED', 5: 'CURRENCY', 6: 'FLOAT_32BIT', 7: 'DOUBLE_64BIT', 8: 'DATE_TIME', 9: 'BINARY_DATA', 10: 'TEXT', 11: 'LARGE_BINARY_DATA', 12: 'LARGE_TEXT', 13: 'SUPER_LARGE_VALUE', 14: 'INETEGER_32BIT_UNSIGNED', 15: 'INTEGER_64BIT_SIGNED', 16: 'GUID', 17: 'INTEGER_16BIT_UNSIGNED'}
	rec = ese_table.get_record(ese_record_num)
	col_type = rec.get_column_type(column_number)
	col_data = rec.get_value_data(column_number)
	#print "rec:%s  col:%s type:%s %s" % (ese_record_num, column_number, col_type, ese_column_types[col_type])
	if col_type == pyesedb.column_types.BINARY_DATA:
		col_data = "" if not col_data else col_data.encode("HEX")
	elif col_type == pyesedb.column_types.BOOLEAN:
		col_data = struct.unpack('?',col_data)[0]
	elif col_type == pyesedb.column_types.CURRENCY:
		pass
	elif col_type == pyesedb.column_types.DATE_TIME:
		pass
	elif col_type == pyesedb.column_types.DOUBLE_64BIT:
		col_data = 0 if not col_data else struct.unpack('d',col_data)[0]
	elif col_type == pyesedb.column_types.FLOAT_32BIT:
		col_data = 0.0 if not col_data else struct.unpack('f',col_data)[0]
	elif col_type == pyesedb.column_types.GUID:
		col_data = str(uuid.UUID(col_data.encode('hex')))    
	elif col_type == pyesedb.column_types.INTEGER_16BIT_SIGNED:
		col_data = 0 if not col_data else struct.unpack('h',col_data)[0]
	elif col_type == pyesedb.column_types.INTEGER_16BIT_UNSIGNED:
		col_data = 0 if not col_data else struct.unpack('H',col_data)[0]
	elif col_type == pyesedb.column_types.INTEGER_32BIT_SIGNED:
		col_data =  0 if not col_data else struct.unpack('i',col_data)[0]
	elif col_type == pyesedb.column_types.INTEGER_32BIT_UNSIGNED:
		col_data = 0 if not col_data else struct.unpack('I',col_data)[0]
	elif col_type == pyesedb.column_types.INTEGER_64BIT_SIGNED:
		col_data = 0 if not col_data else struct.unpack('q',col_data)[0]
	elif col_type == pyesedb.column_types.INTEGER_8BIT_UNSIGNED:
		col_data = 0 if not col_data else struct.unpack('B',col_data)[0]
	elif col_type == pyesedb.column_types.LARGE_BINARY_DATA:
		col_data = "" if not col_data else col_data.encode("HEX")
	elif col_type == pyesedb.column_types.LARGE_TEXT:
		col_data = blob_to_string(col_data)
	elif col_type == pyesedb.column_types.NULL:
		pass
	elif col_type == pyesedb.column_types.SUPER_LARGE_VALUE:
		col_data = "" if not col_data else col_data.encode("HEX")
	elif col_type == pyesedb.column_types.TEXT:
		col_data = blob_to_string(col_data)       
	if col_data==None:
		col_data = "Empty"
	return col_data


ads = (x for x in ["Mark Baggett and Don Williams wrote this program in 3 days. Coding in Python is easy.   Check out SANS Automating Infosec with Python SEC573 to learn to write your own forensics tools.",
       "To learn how SRUM and other artifacts can enhance your forensics investigations check out SANS Windows Forensics FOR408",
       "This program uses the function BinarySIDtoStringSID from the GRR code base to convert binary data into a user SID and relies heavily on the CoreSecurity Impacket ESE module. This works because of them.  Check them out!",
       "Yogesh Khatri's paper at https://files.sans.org/summit/Digital_Forensics_and_Incident_Response_Summit_2015/PDFs/Windows8SRUMForensicsYogeshKhatri.pdf was essential in the creation of this tool.",
       "By modifying the template file you have control of what ends up in the analyzed results.  Try creating an alternate template and passing it with the --XLSX_TEMPLATE option.",
       "This program was written by Twitter:@markbaggett and @donaldjwilliam5 because @ovie said so.",
       "You could analyze other ESE format databases with this by creating a new template.  The program ese_template will get you started genereating templates.",
       ])

parser = argparse.ArgumentParser(description="Given an SRUM database it will create an XLS spreadsheet with analysis of the data in the database.")
parser.add_argument("--ESE_INFILE", help ="Specify the ESE (.dat) file to analyze. Provide a valid path to the file.")
parser.add_argument("--XLSX_OUTFILE", default="SRUM_DUMP_OUTPUT.xlsx", help="Full path to the XLS file that will be created.")
parser.add_argument("--XLSX_TEMPLATE" ,help = "The Excel Template that specifies what data to extract from the srum database. You can create templates with ese_template.py.")
parser.add_argument("--REG_HIVE", dest="reghive", help = "If a registry hive is provided then the names of the network profiles will be resolved.")
parser.add_argument("--quiet",help = "Supress unneeded output messages.",action="store_true")

options = parser.parse_args()

if not options.ESE_INFILE:
    options.ESE_INFILE = raw_input(r"What is the path to the SRUDB.DAT file? (Ex: \image-mount-point\Windows\system32\sru\srudb.dat) : ")
    options.XLSX_OUTFILE = raw_input(r"What is my output file name (including path) (Ex: \users\me\Desktop\resultx.xlsx) : ")
    options.XLSX_TEMPLATE = raw_input("What XLS Template should I use? (Press enter for the Default SRUM_TEMPLATE.XLSX) : ")
    options.reghive = raw_input("What is the path to the SOFTWARE registry hive so I can resolve Network Names? (Press enter to skip resolution) : ")

if not options.XLSX_TEMPLATE:
    options.XLSX_TEMPLATE = "SRUM_TEMPLATE.xlsx"

if options.reghive:
    interface_table = load_interfaces(options.reghive)

warnings.simplefilter("ignore")
try:
    ese_db = pyesedb.open(options.ESE_INFILE)
except Exception as e:
    print "I could not open the specified SRUM file. Check your path and file name."
    print "Error : ", str(e)
    sys.exit(1)

try:
    template_wb = openpyxl.load_workbook(filename=options.XLSX_TEMPLATE, read_only=True)
except Exception as e:
    print "I could not open the specified template file %s. Check your path and file name." % (options.XLSX_TEMPLATE)
    print "Error : ", str(e)
    sys.exit(1)


id_table = load_lookups(ese_db)
target_wb = openpyxl.Workbook()
sheets = template_wb.get_sheet_names()
for each_sheet in sheets[4:]:
    #open the first sheet in the template
    template_sheet = template_wb.get_sheet_by_name(each_sheet)
    #retieve the name of the ESE table to populate the sheet with from A1
    ese_template_table = template_sheet.cell("A1").value
    #retrieve the names of the ESE table columns and cell styles from row 2 and format commands from row 3 
    ese_template_fields = []
    ese_template_formats = []
    ese_template_styles = []
    for eachcolumn in range(1,template_sheet.max_column+1):
        field_name = template_sheet.cell(row = 2, column = eachcolumn).value
        if field_name == None:
            break
        field_style = template_sheet.cell(row = 2, column = eachcolumn).style 
        format_cmd = template_sheet.cell(row = 3, column = eachcolumn).value
        ese_template_formats.append(format_cmd)
        ese_template_styles.append(field_style)
        ese_template_fields.append(field_name.strip())
    #Now open the specified table in the ESE database for this sheet
    try:
        ese_table = ese_db.get_table_by_name(ese_template_table)
    except Exception as e:
        print "Unable to find table",ese_template_table
        print "Error: "+str(e)
        continue
    #Now create the worksheet in the new xls file with the same name as the template
    print "Creating Sheet "+each_sheet
    if not options.quiet:
        try:
            ad = ads.next()
        except:
            ad = "Thanks for using srum_dump!"
        print "While you wait, did you know ...\n"+ad+"\n"

    xls_sheet = target_wb.create_sheet(title=each_sheet)
    #Now copy the header values and header formats from the template to the new worksheet
    header_row = []
    for eachcolumn in range(1,len(ese_template_fields)+1):
        cell_value = template_sheet.cell(row = 4, column = eachcolumn ).value
        cell_style = template_sheet.cell(row = 4, column = eachcolumn).style
        new_cell = WriteOnlyCell(xls_sheet, value=cell_value)
        new_cell.style = cell_style
        header_row.append(new_cell)
    xls_sheet.append(header_row)
    #Until we get an empty row retrieve the rows from the ESE table and process them
    column_lookup = dict([(x.name,index) for index,x in enumerate(ese_table.columns)])
    for ese_row_num in range(ese_table.number_of_records):
        #The row is retrieved now use the template to figure out which ones you want and format them
        xls_row = []
        for eachcolumn,eachformat,eachstyle in zip(ese_template_fields,ese_template_formats,ese_template_styles):
            if eachcolumn == "#XLS_COLUMN#":
                val = eachformat
            else:
                val = smart_retrieve(ese_table, ese_row_num, column_lookup[eachcolumn])
                if val=="Error":
                    val = "WARNING: Invalid Column Name " + eachcolumn+ " - Try one of these:"+str(ese_template_fields) + str(eachcolumn in ese_template_fields)
                elif val==None:
                    val="None"
                elif eachformat == None:
                    pass
                elif eachformat.startswith("OLE:"):
                    val = ole_timestamp(val, eachformat[4:])
                elif eachformat.startswith("FILE:"):
                    val = file_timestamp(val,eachformat[5:])
                elif eachformat.lower() == "lookup_id":
                    val = id_table[val]
                elif eachformat.lower() == "md5":
                    val = hashlib.md5(str(val)).hexdigest()
                elif eachformat.lower() == "sha1":
                    val = hashlib.sha1(str(val)).hexdigest()
                elif eachformat.lower() == "sha256":
                    val = hashlib.sha256(str(val)).hexdigest()
                elif eachformat.lower() == "base16":
                    if type(val)=="<type 'int'>":
                        val = hex(val)
                    else:
                        val = str(val).encode("hex")
                elif eachformat.lower() == "base2":
                    if type(val)==int:
                        val = bin(val)
                    else:
                        try:
                            val = int(str(val),2)
                        except :
                            val = "Warning: Unable to convert value %s to binary." % (val)
                elif eachformat.lower() == "interface_id" and options.reghive:
                    val = interface_table.get(str(val),"")
                elif eachformat.lower() == "interface_id" and not options.reghive:
                    pass
                else:
                    val =  "WARNING: I'm not sure what to do with the format command %s.  It was ignored." % (eachformat)
                new_cell = WriteOnlyCell(xls_sheet, value=val)
                new_cell.style = eachstyle
                #print dir(new_cell.style.font)
                xls_row.append(new_cell)
        xls_sheet.append(xls_row)

firstsheet=target_wb.get_sheet_by_name("Sheet")
target_wb.remove_sheet(firstsheet)
try:
    target_wb.save(options.XLSX_OUTFILE)
except Exception as e:
    print "I was unable to write the output file.  Do you have an old version open?  If not this is probably a path or permissions issue."
    print "Error : ", str(e)



